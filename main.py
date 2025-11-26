"""
Graviton Backend API
FastAPI server for Aadhaar masking and document management
In-memory processing - no files stored on disk
With time-based HMAC authentication
"""

from fastapi import FastAPI, File, UploadFile, HTTPException, Header, Depends, Form, Request, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, HTMLResponse, RedirectResponse
from pydantic import BaseModel
from datetime import datetime, timedelta
import base64
import cv2
import numpy as np
import hmac
import hashlib
import time
import secrets
import json
import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Import the Aadhaar processor
from aadhaar_processor import process_single_image

# ==================== AUTH CONFIGURATION ====================
import os

SECRET_KEY = os.getenv("SECRET_KEY", "GRAVITON_AADHAAR_SECURE_2024")
TOKEN_VALIDITY_MINUTES = 5

# Fixed credentials
ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "gravitonadmin")

# ==================== SESSION MANAGEMENT ====================
active_sessions = {}  # session_id -> username

# ==================== IN-MEMORY LOG STORAGE ====================
LOG_FILE = "request_logs.json"
LOG_RETENTION_DAYS = 7
request_logs = []
MAX_LOGS = 1000  # Increased from 100 for longer retention

def load_logs():
    """Load logs from file on startup"""
    global request_logs
    if os.path.exists(LOG_FILE):
        try:
            with open(LOG_FILE, 'r', encoding='utf-8') as f:
                request_logs = json.load(f)
            print(f"✅ Loaded {len(request_logs)} existing logs from {LOG_FILE}")
            # Clean old logs on startup
            cleanup_old_logs()
        except Exception as e:
            print(f"⚠️ Could not load logs: {e}")
            request_logs = []
    else:
        print("📝 No existing log file found, starting fresh")
        request_logs = []

def save_logs():
    """Save logs to file"""
    try:
        with open(LOG_FILE, 'w', encoding='utf-8') as f:
            json.dump(request_logs, f, indent=2)
    except Exception as e:
        print(f"⚠️ Could not save logs: {e}")

def cleanup_old_logs():
    """Remove logs older than LOG_RETENTION_DAYS"""
    global request_logs
    cutoff_date = datetime.now() - timedelta(days=LOG_RETENTION_DAYS)
    
    original_count = len(request_logs)
    request_logs = [
        log for log in request_logs
        if datetime.fromisoformat(log['timestamp']) > cutoff_date
    ]
    
    removed_count = original_count - len(request_logs)
    if removed_count > 0:
        print(f"🗑️ Cleaned up {removed_count} logs older than {LOG_RETENTION_DAYS} days")
        save_logs()

# ==================== REQUEST MODELS ====================
class LoginRequest(BaseModel):
    username: str
    password: str

# ==================== AUTH FUNCTIONS ====================
def generate_auth_token(timestamp: int = None) -> str:
    """Generate authentication token using HMAC-SHA256"""
    if timestamp is None:
        timestamp = int(time.time())
    
    message = str(timestamp).encode('utf-8')
    signature = hmac.new(
        SECRET_KEY.encode('utf-8'),
        message,
        hashlib.sha256
    ).hexdigest()
    
    token = f"{timestamp}|{signature}"
    return base64.b64encode(token.encode('utf-8')).decode('utf-8')

def verify_auth_token(token: str) -> bool:
    """Verify the authentication token"""
    try:
        decoded = base64.b64decode(token.encode('utf-8')).decode('utf-8')
        parts = decoded.split('|')
        if len(parts) != 2:
            return False
        
        timestamp_str, received_signature = parts
        timestamp = int(timestamp_str)
        
        current_time = int(time.time())
        time_diff = current_time - timestamp
        
        if time_diff < 0 or time_diff > (TOKEN_VALIDITY_MINUTES * 60):
            return False
        
        message = timestamp_str.encode('utf-8')
        expected_signature = hmac.new(
            SECRET_KEY.encode('utf-8'),
            message,
            hashlib.sha256
        ).hexdigest()
        
        return hmac.compare_digest(received_signature, expected_signature)
    
    except Exception:
        return False

async def verify_authorization(authorization: str = Header(None)):
    """Dependency to verify authorization header"""
    if not authorization:
        raise HTTPException(
            status_code=401,
            detail="Missing Authorization header"
        )
    
    token = authorization
    if authorization.startswith("Bearer "):
        token = authorization.replace("Bearer ", "")
    
    if not verify_auth_token(token):
        raise HTTPException(
            status_code=401,
            detail="Invalid or expired authorization token"
        )
    
    return True

app = FastAPI(title="Graviton API", version="2.0.0")

# CORS Configuration
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==================== STARTUP EVENT ====================
@app.on_event("startup")
async def startup_event():
    """Runs once when the server starts"""
    from aadhaar_processor import model, MODEL_PATH
    
    print("\n" + "="*60)
    print("🎬 SERVER STARTUP")
    print("="*60)
    
    if model is not None:
        print(f"✅ YOLO Model loaded successfully")
        print(f"📁 Model path: {MODEL_PATH}")
        print(f"🔄 Model will be REUSED for all requests (no reload)")
    else:
        print("⚠️  WARNING: YOLO Model not loaded!")
        print("   Image processing will fail.")
    
    print("="*60 + "\n")
    
    # Load existing logs
    load_logs()

@app.get("/")
async def root():
    return {
        "message": "Graviton API is running (In-Memory Processing)",
        "version": "2.0.0",
        "description": "Secure API with username/password + time-based HMAC authentication",
        "endpoints": {
            "login": "/api/auth/token - Login to get token (POST)",
            "upload": "/api/aadhaar/upload - Upload image (requires auth)",
            "admin": "/admin/logs - View request logs (requires auth)",
            "health": "/health - Health check (public)"
        },
        "status_codes": {
            "200": "Success - Aadhaar detected and masking applied",
            "422": "Unprocessable Entity - No Aadhaar detected",
            "401": "Unauthorized - Invalid or missing auth token",
            "500": "Internal Server Error"
        }
    }

@app.get("/health")
async def health_check():
    """Health check endpoint with worker statistics"""
    import psutil
    
    # Get worker info from environment if running under Gunicorn
    worker_id = os.getenv("WORKER_ID", "N/A")
    
    # System stats
    cpu_percent = psutil.cpu_percent(interval=0.1)
    memory = psutil.virtual_memory()
    
    return {
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "server_info": {
            "worker_configuration": "5 workers (Gunicorn + Uvicorn)",
            "max_concurrent_requests": 5,
            "current_worker_id": worker_id
        },
        "system_stats": {
            "cpu_usage_percent": cpu_percent,
            "memory_total_gb": round(memory.total / (1024**3), 2),
            "memory_used_gb": round(memory.used / (1024**3), 2),
            "memory_percent": memory.percent
        },
        "request_logs": {
            "total_requests_logged": len(request_logs),
            "max_log_capacity": MAX_LOGS
        }
    }


@app.post("/api/auth/token")
async def login_and_get_token(credentials: LoginRequest):
    """Authenticate with username/password and get authentication token"""
    if credentials.username != ADMIN_USERNAME or credentials.password != ADMIN_PASSWORD:
        raise HTTPException(
            status_code=401,
            detail="Invalid username or password"
        )
    
    token = generate_auth_token()
    
    return {
        "success": True,
        "message": "Authentication successful",
        "token": token,
        "validity_minutes": TOKEN_VALIDITY_MINUTES,
        "usage": {
            "header": "Authorization",
            "value": f"Bearer {token}"
        }
    }

@app.post("/api/aadhaar/upload")
async def upload_aadhaar(
    file: UploadFile = File(...),
    include_all_rotations: bool = Form(True),
    authorized: bool = Depends(verify_authorization),
    request_id: str = Header(None, alias="Request-ID")
):
    """Upload an Aadhaar card image for masking"""
    try:
        # Validate file type
        if not file.content_type.startswith('image/'):
            raise HTTPException(status_code=400, detail="File must be an image")
        
        # Read and decode image
        image_bytes = await file.read()
        input_base64 = base64.b64encode(image_bytes).decode('utf-8')
        
        nparr = np.frombuffer(image_bytes, np.uint8)
        image_array = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        
        if image_array is None:
            raise HTTPException(status_code=400, detail="Invalid image file")
        
        # Process the image
        result = process_single_image(image_array=image_array, include_all_rotations=include_all_rotations)
        
        if result is None:
            raise HTTPException(status_code=500, detail="Image processing failed")
        
        extracted_info, masked_image_array, metrics = result
        
        # Extract confidence from extracted_info
        confidence = extracted_info.get("confidence", 0.0)
        
        # Check if masking was applied
        aadhaar_number = extracted_info.get("AADHAR_NUMBER", "")
        info_has_detection = (
            aadhaar_number and 
            aadhaar_number != "Not detected" and 
            aadhaar_number != "XXXX XXXX XXXX"
        )
        
        # Check for black masking pixels
        original_resized = cv2.resize(image_array, (masked_image_array.shape[1], masked_image_array.shape[0]))
        masked_black = np.all(masked_image_array == [0, 0, 0], axis=2)
        original_black = np.all(original_resized == [0, 0, 0], axis=2)
        new_black_pixels = np.sum(masked_black & ~original_black)
        image_was_masked = new_black_pixels > 100
        
        masking_applied = info_has_detection or image_was_masked
        
        # Convert masked image to base64
        success, buffer = cv2.imencode('.jpg', masked_image_array)
        if not success:
            raise HTTPException(status_code=500, detail="Failed to encode masked image")
        
        output_base64 = base64.b64encode(buffer.tobytes()).decode('utf-8')
        
        # Prepare response
        if masking_applied:
            response_data = {
                "masked_output": output_base64,
                "details": {
                    "already_masked_count": 0,
                    "masking_done_count": 1
                }
            }
            status_code = 200
        else:
            response_data = {
                "masked_output": output_base64,
                "details": {
                    "already_masked_count": 0,
                    "masking_done_count": 0
                }
            }
            status_code = 422
        
        # Log the request - only store images for failed masking
        log_entry = {
            "timestamp": datetime.now().isoformat(),
            "request_id": request_id if request_id else "N/A",
            "status_code": status_code,
            "response_body": response_data,
            "confidence": confidence,
            "performance": metrics
        }
        
        # Only store input image if masking failed (status 422)
        if status_code == 422:
            log_entry["input_base64"] = input_base64
        
        request_logs.append(log_entry)
        
        # Clean old logs before checking max limit
        cleanup_old_logs()
        
        # Keep only last MAX_LOGS entries (after cleanup)
        if len(request_logs) > MAX_LOGS:
            request_logs[:] = request_logs[-MAX_LOGS:]
        
        # Save logs to file
        save_logs()
        
        return JSONResponse(content=response_data, status_code=status_code)
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/admin/login", response_class=HTMLResponse)
async def admin_login_page(request: Request):
    """Admin login page"""
    # Check if already logged in
    session_id = request.cookies.get("admin_session")
    if session_id and session_id in active_sessions:
        return RedirectResponse(url="/admin/logs", status_code=302)
    
    html_content = """
    <!DOCTYPE html>
    <html>
    <head>
        <title>Admin Login - Graviton API</title>
        <style>
            body {
                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                margin: 0;
                padding: 0;
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                min-height: 100vh;
                display: flex;
                justify-content: center;
                align-items: center;
            }
            .login-container {
                background: white;
                border-radius: 12px;
                box-shadow: 0 20px 60px rgba(0,0,0,0.3);
                padding: 40px;
                width: 100%;
                max-width: 400px;
            }
            h1 {
                color: #333;
                margin-bottom: 10px;
                text-align: center;
            }
            .subtitle {
                text-align: center;
                color: #666;
                margin-bottom: 30px;
            }
            .form-group {
                margin-bottom: 20px;
            }
            label {
                display: block;
                margin-bottom: 5px;
                color: #333;
                font-weight: 500;
            }
            input[type="text"],
            input[type="password"] {
                width: 100%;
                padding: 12px;
                border: 2px solid #e0e0e0;
                border-radius: 6px;
                font-size: 14px;
                box-sizing: border-box;
                transition: border-color 0.3s;
            }
            input[type="text"]:focus,
            input[type="password"]:focus {
                outline: none;
                border-color: #667eea;
            }
            button {
                width: 100%;
                padding: 12px;
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                color: white;
                border: none;
                border-radius: 6px;
                font-size: 16px;
                font-weight: 600;
                cursor: pointer;
                transition: transform 0.2s;
            }
            button:hover {
                transform: translateY(-2px);
            }
            .error {
                background: #fee;
                color: #c33;
                padding: 10px;
                border-radius: 6px;
                margin-bottom: 20px;
                display: none;
            }
            .logo {
                text-align: center;
                font-size: 48px;
                margin-bottom: 20px;
            }
        </style>
    </head>
    <body>
        <div class="login-container">
            <div class="logo">🔐</div>
            <h1>Admin Login</h1>
            <p class="subtitle">Graviton API Dashboard</p>
            
            <div id="error" class="error"></div>
            
            <form id="loginForm">
                <div class="form-group">
                    <label for="username">Username</label>
                    <input type="text" id="username" name="username" required autofocus>
                </div>
                <div class="form-group">
                    <label for="password">Password</label>
                    <input type="password" id="password" name="password" required>
                </div>
                <button type="submit">Login</button>
            </form>
        </div>
        
        <script>
            document.getElementById('loginForm').addEventListener('submit', async (e) => {
                e.preventDefault();
                
                const username = document.getElementById('username').value;
                const password = document.getElementById('password').value;
                const errorDiv = document.getElementById('error');
                
                try {
                    const response = await fetch('/admin/login', {
                        method: 'POST',
                        headers: {
                            'Content-Type': 'application/x-www-form-urlencoded',
                        },
                        body: `username=${encodeURIComponent(username)}&password=${encodeURIComponent(password)}`
                    });
                    
                    if (response.ok) {
                        window.location.href = '/admin/logs';
                    } else {
                        const data = await response.json();
                        errorDiv.textContent = data.detail || 'Login failed';
                        errorDiv.style.display = 'block';
                    }
                } catch (error) {
                    errorDiv.textContent = 'An error occurred. Please try again.';
                    errorDiv.style.display = 'block';
                }
            });
        </script>
    </body>
    </html>
    """
    return HTMLResponse(content=html_content)

@app.post("/admin/login")
async def admin_login(response: Response, username: str = Form(...), password: str = Form(...)):
    """Process admin login"""
    if username != ADMIN_USERNAME or password != ADMIN_PASSWORD:
        raise HTTPException(
            status_code=401,
            detail="Invalid username or password"
        )
    
    # Create session
    session_id = secrets.token_urlsafe(32)
    active_sessions[session_id] = username
    
    # Set cookie
    response = JSONResponse(content={"success": True, "message": "Login successful"})
    response.set_cookie(
        key="admin_session",
        value=session_id,
        httponly=True,
        max_age=3600,  # 1 hour
        samesite="lax"
    )
    
    return response

@app.get("/admin/logout")
async def admin_logout(request: Request, response: Response):
    """Logout admin"""
    session_id = request.cookies.get("admin_session")
    if session_id and session_id in active_sessions:
        del active_sessions[session_id]
    
    response = RedirectResponse(url="/admin/login", status_code=302)
    response.delete_cookie("admin_session")
    return response

async def verify_admin_session(request: Request):
    """Dependency to verify admin session"""
    session_id = request.cookies.get("admin_session")
    if not session_id or session_id not in active_sessions:
        raise HTTPException(
            status_code=401,
            detail="Not authenticated. Please login at /admin/login"
        )
    return active_sessions[session_id]

@app.get("/admin/logs/download")
async def download_logs_excel(request: Request, username: str = Depends(verify_admin_session)):
    """Download logs as Excel file"""
    from fastapi.responses import StreamingResponse
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Request Logs"
    
    # Define headers
    headers = ["Timestamp", "Request ID", "Status Code", "Masking Done", "Already Masked", "Confidence", 
               "Preprocessing (ms)", "Model Inference (ms)", "Postprocessing (ms)"]
    
    # Add input image column only if there are failed requests
    has_failed_requests = any(log['status_code'] == 422 for log in request_logs)
    if has_failed_requests:
        headers.append("Input Image (Base64)")
    
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data
    for log in request_logs:
        masking_done = log['response_body']['details']['masking_done_count']
        already_masked = log['response_body']['details']['already_masked_count']
        confidence = log.get('confidence', 0.0)
        request_id = log.get('request_id', 'N/A')
        
        # Extract performance metrics
        perf = log.get('performance', {})
        preprocessing_ms = perf.get('3a_preprocessing_ms', 0.0)
        inference_ms = perf.get('3_model_inference_total_ms', 0.0)
        postprocessing_ms = perf.get('4a_postproc_validation_ms', 0.0)
        
        row = [
            log['timestamp'],
            request_id,
            log['status_code'],
            masking_done,
            already_masked,
            round(confidence, 4),
            round(preprocessing_ms, 2),
            round(inference_ms, 2),
            round(postprocessing_ms, 2)
        ]
        
        # Add input image only if present (failed requests)
        if has_failed_requests:
            input_img = log.get('input_base64', '') if log.get('input_base64') else 'N/A'
            row.append(input_img)
        
        ws.append(row)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20
    if has_failed_requests:
        ws.column_dimensions['J'].width = 30
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Return as download
    filename = f"request_logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/admin/logs", response_class=HTMLResponse)
async def admin_logs(request: Request, username: str = Depends(verify_admin_session)):
    """Admin panel to view request logs (requires session authentication)"""
    
    html_content = """
    <!DOCTYPE html>
    <html>
    <head>
        <title>Admin Logs - Graviton API</title>
        <style>
            * {
                margin: 0;
                padding: 0;
                box-sizing: border-box;
            }
            body {
                font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif;
                background: #f5f5f5;
                color: #333;
                line-height: 1.6;
            }
            .container {
                max-width: 1400px;
                margin: 20px auto;
                background: white;
                border: 1px solid #e0e0e0;
                box-shadow: 0 1px 3px rgba(0,0,0,0.1);
            }
            .header {
                padding: 20px 30px;
                border-bottom: 1px solid #e0e0e0;
                display: flex;
                justify-content: space-between;
                align-items: center;
            }
            h1 {
                font-size: 24px;
                font-weight: 600;
                color: #1a1a1a;
            }
            .user-info {
                color: #666;
                font-size: 14px;
            }
            .actions {
                display: flex;
                gap: 15px;
                align-items: center;
            }
            .btn {
                padding: 8px 16px;
                text-decoration: none;
                font-size: 14px;
                font-weight: 500;
                border-radius: 4px;
                transition: all 0.2s;
                display: inline-block;
            }
            .btn-primary {
                background: #2196F3;
                color: white;
            }
            .btn-primary:hover {
                background: #1976D2;
            }
            .btn-secondary {
                color: #666;
            }
            .btn-secondary:hover {
                color: #333;
            }
            .content {
                padding: 30px;
            }
            .stats {
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
                gap: 20px;
                margin-bottom: 30px;
            }
            .stat-card {
                padding: 20px;
                border: 1px solid #e0e0e0;
                border-radius: 4px;
                text-align: center;
            }
            .stat-value {
                font-size: 32px;
                font-weight: 600;
                color: #1a1a1a;
                margin-bottom: 5px;
            }
            .stat-label {
                font-size: 13px;
                color: #666;
                text-transform: uppercase;
                letter-spacing: 0.5px;
            }
            table {
                width: 100%;
                border-collapse: collapse;
                font-size: 14px;
            }
            th {
                background: #fafafa;
                padding: 12px 16px;
                text-align: left;
                font-weight: 600;
                color: #333;
                border-bottom: 2px solid #e0e0e0;
                font-size: 13px;
                text-transform: uppercase;
                letter-spacing: 0.5px;
            }
            td {
                padding: 12px 16px;
                border-bottom: 1px solid #f0f0f0;
            }
            tr:hover {
                background: #fafafa;
            }
            .status-200 {
                color: #4CAF50;
                font-weight: 600;
            }
            .status-422 {
                color: #FF9800;
                font-weight: 600;
            }
            .status-500 {
                color: #F44336;
                font-weight: 600;
            }
            .image-preview {
                width: 60px;
                height: 45px;
                object-fit: cover;
                border: 1px solid #e0e0e0;
                border-radius: 2px;
                cursor: pointer;
            }
            .image-preview:hover {
                opacity: 0.8;
            }
            .badge {
                display: inline-block;
                padding: 4px 8px;
                border-radius: 3px;
                font-size: 12px;
                font-weight: 500;
            }
            .badge-success {
                background: #E8F5E9;
                color: #2E7D32;
            }
            .badge-warning {
                background: #FFF3E0;
                color: #E65100;
            }
            .modal {
                display: none;
                position: fixed;
                z-index: 1000;
                left: 0;
                top: 0;
                width: 100%;
                height: 100%;
                background-color: rgba(0,0,0,0.85);
            }
            .modal-content {
                margin: auto;
                display: block;
                max-width: 90%;
                max-height: 90%;
                margin-top: 50px;
            }
            .close {
                position: absolute;
                top: 15px;
                right: 35px;
                color: #f1f1f1;
                font-size: 40px;
                font-weight: bold;
                cursor: pointer;
            }
            .no-logs {
                text-align: center;
                padding: 60px 20px;
                color: #999;
            }
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <div>
                    <h1>Admin Request Logs</h1>
                    <div class="user-info">Logged in as: <strong>""" + username + """</strong></div>
                </div>
                <div class="actions">
                    <a href="/admin/logs/download" class="btn btn-primary">Download Excel</a>
                    <a href="/admin/logout" class="btn btn-secondary">Logout</a>
                </div>
            </div>
            
            <div class="content">
            <div class="stats">
                <div class="stat-card">
                    <div class="stat-value">""" + str(len(request_logs)) + """</div>
                    <div class="stat-label">Total Requests</div>
                </div>
                <div class="stat-card">
                    <div class="stat-value">""" + str(sum(1 for log in request_logs if log['status_code'] == 200)) + """</div>
                    <div class="stat-label">Success</div>
                </div>
                <div class="stat-card">
                    <div class="stat-value">""" + str(sum(1 for log in request_logs if log['status_code'] == 422)) + """</div>
                    <div class="stat-label">No Detection</div>
                </div>
            </div>
    """
    
    if not request_logs:
        html_content += """
            <div class="no-logs">
                <h2>No requests logged yet</h2>
                <p>Make some requests to /api/aadhaar/upload to see them here</p>
            </div>
        """
    else:
        html_content += """
            <table>
                <thead>
                    <tr>
                        <th>Timestamp</th>
                        <th>Request ID</th>
                        <th>Status Code</th>
                        <th>Masking Status</th>
                        <th>Confidence</th>
                        <th>Preprocessing (ms)</th>
                        <th>Inference (ms)</th>
                        <th>Postproc (ms)</th>
                        <th>Input Image</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        for log in reversed(request_logs):
            status_class = f"status-{log['status_code']}"
            masking_done = log['response_body']['details']['masking_done_count']
            confidence = log.get('confidence', 0.0)
            request_id = log.get('request_id', 'N/A')
            
            # Extract performance metrics
            perf = log.get('performance', {})
            preprocessing_ms = perf.get('3a_preprocessing_ms', 0.0)
            inference_ms = perf.get('3_model_inference_total_ms', 0.0)
            postprocessing_ms = perf.get('4a_postproc_validation_ms', 0.0)
            
            # Only show input image if masking failed
            input_image_cell = ""
            if log.get('input_base64'):
                input_image_cell = f"""
                    <img src="data:image/jpeg;base64,{log['input_base64'][:100]}..." 
                         class="image-preview" 
                         onclick="showImage('data:image/jpeg;base64,{log['input_base64']}')"
                         alt="Input">
                """
            else:
                input_image_cell = "<span style='color: #999;'>N/A</span>"
            
            html_content += f"""
                <tr>
                    <td>{log['timestamp']}</td>
                    <td>{request_id}</td>
                    <td class="{status_class}">{log['status_code']}</td>
                    <td>
                        <span class="badge {'badge-success' if masking_done == 1 else 'badge-warning'}">
                            {'Success' if masking_done == 1 else 'Failed'}
                        </span>
                    </td>
                    <td>{round(confidence, 4)}</td>
                    <td>{round(preprocessing_ms, 2)}</td>
                    <td>{round(inference_ms, 2)}</td>
                    <td>{round(postprocessing_ms, 2)}</td>
                    <td>{input_image_cell}</td>
                </tr>
            """
        
        html_content += """
                </tbody>
            </table>
            </div>
        """
    
    html_content += """
            <div id="imageModal" class="modal" onclick="this.style.display='none'">
                <span class="close">&times;</span>
                <img class="modal-content" id="modalImage">
            </div>
        </div>
        
        <script>
            function showImage(src) {
                document.getElementById('modalImage').src = src;
                document.getElementById('imageModal').style.display = 'block';
            }
        </script>
    </body>
    </html>
    """
    
    return HTMLResponse(content=html_content)

@app.get("/health")
async def health_check():
    return {
        "status": "healthy",
        "timestamp": datetime.now().isoformat()
    }

if __name__ == "__main__":
    import uvicorn
    import socket
    
    hostname = socket.gethostname()
    local_ip = socket.gethostbyname(hostname)
    
    print("\n" + "="*60)
    print("🚀 Graviton Backend Server Starting...")
    print("="*60)
    print(f"📍 Local Access:    http://localhost:8000")
    print(f"🌐 Network Access:  http://{local_ip}:8000")
    print(f"📖 API Docs:        http://localhost:8000/docs")
    print(f"👨‍💼 Admin Panel:     http://localhost:8000/admin/logs")
    print("="*60)
    print("✨ In-Memory Processing - No files stored on disk")
    print("🔒 All images processed and returned as base64")
    print(f"🔐 Authentication:  Username/Password + HMAC-SHA256")
    print("="*60)
    print("\n🔑 Authentication Steps:")
    print("1. POST /api/auth/token with credentials:")
    print(f'   {{"username": "{ADMIN_USERNAME}", "password": "***"}}')
    print("2. Receive token (valid for 5 minutes)")
    print("3. Add header: Authorization: Bearer <token>")
    print("4. POST /api/aadhaar/upload - Upload with auth")
    print("5. GET /admin/logs - View request logs")
    print("="*60 + "\n")
    
    uvicorn.run(app, host="0.0.0.0", port=8000)
